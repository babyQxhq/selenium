#coding=utf-8
import openpyxl
from openpyxl.styles import Border,Side,Font
import time
class ParseExcel(object):
    def __init__(self):
        self.workbook=None
        self.excelFile=None
        #设置字体颜色
        self.font=Font(color=None)
        #设置颜色对应RGB
        self.RGBDict={'red':'FFFF3030','green':'FF008B00'}
    def loadWorkBook(self,excelPathAndName):
        #将excel加载到内存中并获取workbook对象
        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e

        self.excelFile=excelPathAndName
        return self.workbook
    def getSheetByName(self,sheetName):
        #通过名字获取sheet对象
        try:
            sheet=self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        #通过索引获取sheet对象
        try:
            sheetName = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet=self.workbook.get_sheet_by_Name(sheetName)
        return sheet
    def getRowsNumber(self,sheet):
        #获取结束行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        #获取结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取开始行号
        return sheet.min_row

    def getStartColsNumber(self, sheet):
        #获取开始列号
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        #获取某行内容用元组存储
        try:
            return list(sheet.rows)[rowNo-1]
        except Exception as e:
            raise e
    def getColumn(self,sheet,colNo):
        # 获取某列内容用元组存储
        try:
            return list(sheet.columns)[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        # 根据单元格所在的位置索引获取单元格值
        #sheet.cell(row=1,column=1).value,表示第一行第一列的值
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate==None and rowNo!=None and colsNo!=None:
            try:
                return sheet.cell(row=rowNo,column=colsNo).value
            except Exception as e:
                raise e

        else:
            raise Exception("没有取到单元格值")

    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        #获取某个单元格的对象，可以根据单元格所在位置的数字索引，也可以根据Excel中坐标及数字索引
        if coordinate!=None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate==None and rowNo!=None and colsNo!=None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise e

        else:
            raise Exception("没有取到单元格值")

    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        #向单元格写入值，content为写入值，style为颜色如red
        if coordinate!=None:
            try:
                sheet.cell(coordinate=coordinate).value=content
                if style!=None:
                    sheet.cell(coordinate=coordinate).font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e



        elif coordinate==None and rowNo!=None and colsNo!=None:
            try:
                sheet.cell(row=rowNo,column=colsNo).value=content
                if style:
                    sheet.cell(row=rowNo,column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e

        else:
            raise Exception("没有取到单元格值")

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        # 写入当前时间
        now=int(time.time())
        timeArray=time.localtime(now)
        currentTime=time.strftime("%Y-%m-%d %H:%M:%S",timeArray)


        if coordinate != None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)


            except Exception as e:
                raise e
        elif coordinate == None and rowNo != None and colsNo != None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e

        else:
            raise Exception("写入时间失败")